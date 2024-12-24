import os
import shutil
import warnings
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from common.log import Log

# 忽略特定的警告
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
# 设置显示选项
pd.set_option('display.max_rows', None)        # 显示所有行
pd.set_option('display.max_columns', None)     # 显示所有列
pd.set_option('display.max_colwidth', None)    # 显示完整的列内容
pd.set_option('display.width', None)           # 宽屏显示
pd.set_option('display.multi_sparse', False)   # 禁用多层索引换行显示

class Excel:
    def __init__(self, report_file_path, case_file_path=None, case_file_sheet=None, output_file_path=None):
        self.case_file_path = case_file_path
        self.case_file_sheet = case_file_sheet
        self.report_file_path = report_file_path
        self.case_dataframe = None  # 用于缓存用例数据
        self.report_excel = None  # 用于缓存报告文件
        self.output_file_path = output_file_path  # 输出文件路径
        self.log = Log()  # 日志器对象

    def load_case_data(self, fillna_value=''):
        """加载用例数据到DataFrame中并填充缺失值"""
        if self.case_file_path is None:
            self.log.info("不需要用例数据")
            return

        if self.case_dataframe is None:
            self.log.info("开始加载用例数据")
            try:
                self.case_dataframe = pd.read_excel(self.case_file_path, sheet_name=self.case_file_sheet)
                self.fillna(df=self.case_dataframe, value=fillna_value)
                self.log.info("用例数据加载完成")
            except Exception as e:
                self.log.error(f"用例加载发生错误: {e}")

    def load_report_data(self):
        """加载整个报告Excel文件至缓存"""
        if self.report_excel is None:
            self.log.info("开始加载报告数据")
            try:
                self.report_excel = pd.ExcelFile(self.report_file_path)
                self.log.info("报告数据加载完成")
            except Exception as e:
                self.log.error(f"报告加载发生错误: {e}")

    def read_caseData(self, module_name, case_name):
        """根据模块名和用例名筛选数据并转换为列表"""
        self.log.info(f"开始读取用例数据: 模块名-{module_name}, 用例名称-{case_name}")
        self.load_case_data()
        df = self.case_dataframe
        filtered_df = df[(df['所属模块'] == module_name) & (df['用例名称(关键字)'] == case_name)]
        if not filtered_df.empty:
            data_list = filtered_df["数据"].values[0].split(',')
            result = [item.strip() for item in data_list]
            self.log.info(f"用例数据读取成功: {result}")
            return result
        else:
            self.log.warning(f"没有找到工作表模块：{module_name} 对应的测试用例数据：{case_name}")
            return []

    def read_report_sheetNames(self):
        """读取报告文件的所有工作表名称并转换为列表"""
        self.log.info("开始读取报告中的所有工作表名称")
        self.load_report_data()
        sheet_names = self.report_excel.sheet_names
        self.log.info(f"报告中的工作表名称读取成功: {sheet_names}")
        return sheet_names

    def read_report_headers(self, sheet_name):
        """读取指定工作表的首行表头并转换为列表"""
        self.log.info(f"开始读取工作表 {sheet_name} 的首行表头")
        df = pd.read_excel(self.report_file_path, sheet_name=sheet_name, nrows=1)
        headers = df.columns.tolist()
        self.log.info(f"工作表 {sheet_name} 的首行表头读取成功: {headers}")
        return headers

    def get_ApplicationInfo(self, report_file_path=None, sheet_name=None, row_indices=None,
                            date_format="%Y-%m-%d %H:%M:%S"):
        """
        获取报告应用信息内容，并将日期时间格式化为指定格式。

        :param report_file_path: Excel 报告文件路径。如果未提供，则使用初始化时提供的 report_file_path。
        :param sheet_name: 工作表名称。如果未提供，则为 应用信息。
        :param row_indices: 行索引列表（基于1）默认为 [1, 2]。
        :param date_format: 日期时间格式，默认是 "%Y-%m-%d %H:%M:%S"。
        :return: 包含指定行内容的字典。
        """
        if row_indices is None:
            row_indices = [1, 2]
        if report_file_path is None:
            report_file_path = self.report_file_path
        if sheet_name is None:
            sheet_name = "应用信息"

        # 加载工作簿并选择指定的工作表
        try:
            wb = load_workbook(filename=report_file_path, data_only=True)  # 使用 data_only=True 来确保只读取计算后的值
            ws = wb[sheet_name]

            # 获取指定行的内容
            selected_rows = []
            for idx in row_indices:
                row_data = [cell.value for cell in ws[idx]]  # 行索引从1开始
                selected_rows.append(row_data)

            # 格式化日期时间字段并创建字典
            Info_dict = {
                key: value.strftime(date_format) if isinstance(value, datetime) else str(value)
                for key, value in zip(selected_rows[0], selected_rows[1])
            }
            self.log.info(f"获取报告应用信息成功: {Info_dict}")
            print("获取报告应用信息成功")
            return Info_dict
        except Exception as e:
            self.log.error(f"获取报告应用信息失败: {e}")
            print("获取报告应用信息失败")
            return None

    def create_report_result(self):
        """创建或打开标记副本"""
        if not os.path.exists(self.output_file_path):
            # 如果副本不存在，复制原始文件作为副本
            shutil.copyfile(self.report_file_path, self.output_file_path)
            self.log.info(f"已创建新的标记副本 {self.output_file_path}")
        else:
            self.log.info(f"正在使用现有的标记副本 {self.output_file_path}")

    def get_ReportSheet(self, sheet_name, fillna_value=''):
        """从报告文件中加载指定工作表的数据并返回为DataFrame"""
        try:
            df = pd.read_excel(self.report_file_path, sheet_name=sheet_name)

            # 统一使用空字符串填充所有类型的缺失值
            df.fillna(str(fillna_value), inplace=True)

            return df
        except Exception as e:
            self.log.error(f"加载工作表 {sheet_name} 发生错误: {e}")
            return None

    def mark_cells_in_sheet(self, failed_cells, sheet_name):
        """为指定工作表中的单元格设置背景颜色并保存"""
        wb = load_workbook(self.output_file_path)
        ws = wb[sheet_name]

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for row_index, col_index in failed_cells:
            cell = ws.cell(row=row_index, column=col_index)
            cell.fill = yellow_fill

        # Save the workbook to the output file path
        wb.save(self.output_file_path)
        self.log.info(f"已保存至 {self.output_file_path}")

    def find_matching_rows(self, json_dict, keys_to_match, df):
        """根据传入的Python字典查找匹配行"""
        conditions = [f"({key} == '{json_dict.get(key)}')" for key in keys_to_match if key in json_dict]
        query_string = ' and '.join(conditions)

        try:
            matching_rows = df.query(query_string)
            self.log.info(f"找到 {len(matching_rows)} 行匹配")
            return matching_rows if not matching_rows.empty else pd.DataFrame(columns=df.columns)  # 确保返回的DataFrame有列名
        except Exception as e:
            self.log.error(f"查询过程中发生错误: {e}")
            return pd.DataFrame(columns=df.columns)  # 返回空DataFrame但保留列结构

    def verify_other_fields(self, matching_row, json_dict):
        """验证匹配行中的其他字段是否一致"""
        failed_cells = []

        # 检查是否找到了匹配行
        if matching_row.empty:
            self.log.warning("没有找到匹配行")
            return failed_cells

        # 确保我们有一个有效的DataFrame或Series
        if isinstance(matching_row, pd.Series):
            matching_row = matching_row.to_frame().transpose()

        # 验证其他字段
        for key, value in json_dict.items():
            if key in matching_row.columns:
                report_value = matching_row[key].iloc[0]
                if report_value != value:
                    # 记录不匹配的详细信息
                    self.log.error(
                        f"测试失败：{key} 不匹配\n"
                        f"  API提供的值: {value}\n"
                        f"  报告中的值: {report_value}"
                    )
                    failed_cells.append((matching_row.index[0] + 2, matching_row.columns.get_loc(key) + 1))

        return failed_cells

    def fillna(self, df, value='', method=None, axis=0, inplace=False, limit=None):
        """
        使用指定的方法填充DataFrame中的缺失值。

        参数:
        - df: 要填充的DataFrame。
        - value: 用来填充缺失值的具体值，默认为空字符串。
        - method: {'backfill', 'bfill', 'pad', 'ffill'}，指定填充方法。
        - axis: {0 or 'index', 1 or 'columns'}，沿哪个轴填充。
        - inplace: 如果为True，则直接修改原始DataFrame，否则返回新的DataFrame。
        - limit: 最大填充次数。
        """
        if df is None:
            self.log.error("无法识别的目标DataFrame")
            return

        try:
            filled_df = df.fillna(value=value, method=method, axis=axis, limit=limit)
            if inplace:
                df.update(filled_df)  # 更新原始DataFrame
            else:
                return filled_df  # 返回新的DataFrame
            self.log.info("缺失值填充完成")
        except Exception as e:
            self.log.error(f"缺失值填充发生错误: {e}")

        # 确保填充后数据类型的正确性
        for col in df.columns:
            if pd.api.types.is_numeric_dtype(df[col]):
                df[col] = pd.to_numeric(df[col], errors='coerce')  # 强制转换数值型列
            elif pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = pd.to_datetime(df[col], errors='coerce')  # 强制转换日期时间型列

    def close(self):
        """清理资源"""
        self.log.info("开始清理Excel实例资源")
        if self.report_excel:
            self.report_excel.close()
            self.report_excel = None
        if self.case_dataframe is not None:
            del self.case_dataframe
            self.case_dataframe = None
        self.log.info("Excel实例资源清理完成")


def str_to_list(str_val):
    """将字符串转换为列表，去除空格"""
    return [item.strip() for item in str_val.split(',')]


if __name__ == '__main__':
    excel_instance = Excel(report_file_path=r"D:\sca_load\【应用报告】Report_Test7214@1.0-20241212162739.xlsx")
    print(excel_instance.get_ReportSheet("检出组件信息"))
